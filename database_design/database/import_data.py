import openpyxl
from openpyxl.worksheet.filters import (
    FilterColumn,
    )

book = openpyxl.load_workbook('./materials.xlsx')
sheet = book.active
max_row = sheet.max_row



manufacturer_col = 1


cells = sheet['A1': 'D5']
#print(cells)
#for c1, c2, c3, c4 in cells:
#    print(f"{c1.value} {c2.value} {c3.value} {c4.value}")


# Loop will print all columns name
for i in range(2, max_row):
    cell_obj = sheet.cell(row=i, column=2)
    #print(cell_obj)
#    print(cell_obj.value)


def parse_cells(cells):
    key_array = []
    data_array = []
    index = 0
    for row in cells:
        #print(row)
        key_val = row[0].value
        # If entry not already in array

        if not key_val in key_array and key_val is not None:
            key_array.append(row[0].value)
            # loop over tuple values
            row_data = [index]
            for cell in row:
                print(cell.value)
                row_data.append(cell.value)
            data_array.append(row_data)
            index=index + 1
                
    print(data_array)

    # print(cells[0])
    #print(len(cells[0]))

def get_manufacturer_names_from_excel():
    manufacturer_names = []
    
    manufacturer_name_col = 2
    for i in range(manufacturer_name_col, max_row):
        cell_obj = sheet.cell(row=i, column=2)
        #print(cell_obj.value)
        if not cell_obj.value in manufacturer_names:
            manufacturer_names.append(cell_obj.value)
    return manufacturer_names





manufacturer_cells = sheet['B2:B61']
materials_cells = sheet['C2': 'E61']

parse_cells(manufacturer_cells)